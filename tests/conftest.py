"""
Fixtures compartidos para tests de ingesta PDP.

Genera archivos de prueba pequeños para cada parser
sin depender de los datos reales de INEGI.
"""

import csv
import io
import os
import sqlite3
import struct
import tempfile
import zipfile
from pathlib import Path

import openpyxl
import pytest


# ── Rutas ────────────────────────────────────────────────────
FIXTURES_DIR = Path(__file__).parent / 'fixtures'
FIXTURES_DIR.mkdir(exist_ok=True)


# ── Helpers ──────────────────────────────────────────────────

def _make_dbf(path: Path, records: list[dict], field_defs: list[tuple]):
    """
    Crea un archivo DBF dBase III simple.
    
    field_defs: [(nombre, tipo, longitud, decimales), ...]
      tipo: 'C' = caracter, 'N' = numérico
    records: [{campo: valor, ...}, ...]
    """
    num_records = len(records)
    header_size = 32 + 32 * len(field_defs) + 1
    record_size = sum(f[2] for f in field_defs) + 1  # +1 delete flag
    
    with open(path, 'wb') as f:
        # Header
        f.write(b'\x03')  # version dBase III
        f.write(struct.pack('B', 2024 - 1900))  # year
        f.write(struct.pack('B', 6))  # month
        f.write(struct.pack('B', 15))  # day
        f.write(struct.pack('<I', num_records))
        f.write(struct.pack('<H', header_size))
        f.write(struct.pack('<H', record_size))
        f.write(b'\x00' * 20)  # reserved
        
        # Field descriptors
        for name, ftype, length, decimal in field_defs:
            name_bytes = name.encode('ascii').ljust(11)[:11]
            f.write(name_bytes)
            f.write(ftype.encode('ascii'))
            f.write(b'\x00' * 4)
            f.write(struct.pack('B', length))
            f.write(struct.pack('B', decimal))
            f.write(b'\x00' * 14)
        
        f.write(b'\x0D')  # header terminator
        
        # Records
        for rec in records:
            f.write(b'\x20')  # not deleted
            for name, ftype, length, decimal in field_defs:
                val = rec.get(name, '')
                if ftype == 'N':
                    if val is None or val == '':
                        encoded = b' ' * length
                    else:
                        s = f'{float(val):>.{decimal}f}' if decimal > 0 else str(int(val))
                        # Left-pad with spaces to exact length
                        encoded = s.rjust(length).encode('ascii')
                        if len(encoded) > length:
                            encoded = encoded[:length]
                        elif len(encoded) < length:
                            encoded = encoded.rjust(length)
                else:
                    if val is None:
                        encoded = b' ' * length
                    else:
                        s = str(val)
                        encoded = s.encode('cp850', errors='replace').ljust(length)[:length]
                assert len(encoded) == length, f'Field {name}: expected {length} bytes, got {len(encoded)} ({encoded!r})'
                f.write(encoded)


# ── Fixtures: DBF ────────────────────────────────────────────

@pytest.fixture(scope='session')
def dbf_sample_path() -> Path:
    """DBF pequeño con 3 campos y 3 registros."""
    path = FIXTURES_DIR / 'test_sample.dbf'
    if not path.exists():
        _make_dbf(path, [
            {'CODIGO': '001', 'NOMBRE': 'Jalisco', 'VALOR': 1500.50},
            {'CODIGO': '002', 'NOMBRE': 'Veracruz', 'VALOR': 890.25},
            {'CODIGO': '003', 'NOMBRE': 'Oaxaca', 'VALOR': 2340.00},
        ], [
            ('CODIGO', 'C', 5, 0),
            ('NOMBRE', 'C', 20, 0),
            ('VALOR', 'N', 8, 2),
        ])
    return path


@pytest.fixture(scope='session')
def dbf_sample_cp850_path() -> Path:
    """DBF con caracteres CP850 (acentos, eñes)."""
    path = FIXTURES_DIR / 'test_sample_cp850.dbf'
    if not path.exists():
        _make_dbf(path, [
            {'CODIGO': '014', 'NOMBRE': 'México', 'VALOR': 5000},
            {'CODIGO': '015', 'NOMBRE': 'Michoacán', 'VALOR': 3200},
        ], [
            ('CODIGO', 'C', 5, 0),
            ('NOMBRE', 'C', 20, 0),
            ('VALOR', 'N', 6, 0),
        ])
    return path


@pytest.fixture(scope='session')
def dbf_zip_path(dbf_sample_path) -> Path:
    """ZIP conteniendo un DBF (simula estructura INEGI 2017+)."""
    path = FIXTURES_DIR / 'test_dbf_in_zip.dbf.zip'
    if not path.exists():
        with zipfile.ZipFile(path, 'w') as zf:
            zf.write(dbf_sample_path, 'Bases_Datos/TEST_SAMPLE.DBF')
    return path


@pytest.fixture(scope='session')
def dbf_zip_2011_path(dbf_sample_path) -> Path:
    """ZIP con estructura INEGI 2011."""
    path = FIXTURES_DIR / 'test_dbf_2011.dbf.zip'
    if not path.exists():
        with zipfile.ZipFile(path, 'w') as zf:
            zf.write(dbf_sample_path, 'Base de datos/TEST_SAMPLE.DBF')
    return path


@pytest.fixture(scope='session')
def dbf_zip_2013_path(dbf_sample_path) -> Path:
    """ZIP con estructura INEGI 2013."""
    path = FIXTURES_DIR / 'test_dbf_2013.dbf.zip'
    if not path.exists():
        with zipfile.ZipFile(path, 'w') as zf:
            zf.write(dbf_sample_path, 'modulo_ejemplo_cngmd2013_dbf/Bases_de_datos/TEST_SAMPLE.DBF')
    return path


# ── Fixtures: CSV ────────────────────────────────────────────

@pytest.fixture(scope='session')
def csv_sample_path() -> Path:
    """CSV pequeño con encoding CP850."""
    path = FIXTURES_DIR / 'test_sample.csv'
    if not path.exists():
        with open(path, 'w', encoding='cp850', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['folio', 'entidad', 'municipio', 'valor'])
            writer.writerow(['1001', 'Jalisco', 'Guadalajara', '95.2'])
            writer.writerow(['1002', 'Veracruz', 'Xalapa', '88.5'])
            writer.writerow(['1003', 'Oaxaca', 'Oaxaca', '76.3'])
    return path


# ── Fixtures: XLSX ───────────────────────────────────────────

@pytest.fixture(scope='session')
def xlsx_tematico_path() -> Path:
    """XLSX en formato temático (2021+) con 2 hojas."""
    path = FIXTURES_DIR / 'test_tematico.xlsx'
    if not path.exists():
        wb = openpyxl.Workbook()
        
        # Hoja Índice
        ws_idx = wb.active
        ws_idx.title = 'Índice'
        ws_idx['A1'] = 'INEGI. Censo Nacional de Gobiernos Estatales 2021. Tabulados básicos'
        ws_idx['B2'] = 'Índice'
        ws_idx['B4'] = 1
        ws_idx['C4'] = 'Nombre del cuadro'
        
        # Hoja 1
        ws1 = wb.create_sheet('1')
        ws1['A1'] = 'INEGI. Censo Nacional. Tabulados básicos'
        ws1['A2'] = 'Condición de existencia de disposiciones normativas'
        ws1['A3'] = '2020'
        ws1['A4'] = 'Clave'
        ws1['B4'] = 'Entidad federativa'
        ws1['C4'] = 'Valor'
        ws1['A6'] = 1
        ws1['B6'] = 'Aguascalientes'
        ws1['C6'] = 31
        ws1['A7'] = 2
        ws1['B7'] = 'Baja California'
        ws1['C7'] = 28
        
        wb.save(path)
        wb.close()
    return path


@pytest.fixture(scope='session')
def xlsx_mformat_path() -> Path:
    """XLSX en formato M (2017-2019) con celdas combinadas."""
    path = FIXTURES_DIR / 'test_mformat.xlsx'
    if not path.exists():
        wb = openpyxl.Workbook()
        
        ws_idx = wb.active
        ws_idx.title = 'Índice'
        ws_idx['A1'] = 'INEGI. Censo Nacional. Tabulados básicos'
        ws_idx['B2'] = 'Índice'
        ws_idx['B4'] = 1.1
        ws_idx['C4'] = 'Nombre del cuadro'
        
        ws1 = wb.create_sheet('1.1')
        ws1['A1'] = 'INEGI. Censo Nacional. Tabulados básicos'
        ws1['A2'] = 'Título del cuadro'
        ws1['A3'] = '2016'
        ws1['A4'] = 'Entidad Federativa'
        ws1['B4'] = 'Total'
        ws1['C4'] = 'Hombres'
        ws1['D4'] = 'Mujeres'
        ws1['A6'] = 'Estados Unidos Mexicanos'
        ws1['B6'] = 100
        ws1['C6'] = 55
        ws1['D6'] = 45
        ws1['A7'] = 'Aguascalientes'
        ws1['B7'] = 10
        ws1['C7'] = 5
        ws1['D7'] = 5
        
        wb.save(path)
        wb.close()
    return path


# ── Fixtures: SQLite (para test de cargador) ─────────────────

@pytest.fixture(scope='session')
def sqlite_sample_path(dbf_sample_path, xlsx_tematico_path) -> Path:
    """Base SQLite pequeña con catálogo + registros."""
    path = FIXTURES_DIR / 'test_cargador.db'
    if not path.exists():
        conn = sqlite3.connect(str(path))
        conn.executescript("""
            CREATE TABLE catalogo (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                censo TEXT NOT NULL, nivel TEXT NOT NULL,
                año INTEGER NOT NULL, periodo INTEGER,
                modulo TEXT NOT NULL, tipo_fuente TEXT NOT NULL,
                nombre_tabla TEXT NOT NULL, archivo_origen TEXT,
                total_registros INTEGER DEFAULT 0,
                total_campos INTEGER DEFAULT 0
            );
            CREATE TABLE registro (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                fuente_id INTEGER NOT NULL REFERENCES catalogo(id),
                datos_json TEXT NOT NULL,
                clave_geo TEXT, entidad TEXT, municipio TEXT
            );
        """)
        conn.execute("INSERT INTO catalogo VALUES (1,'CNGMD','municipal',2021,2020,'agua','XLSX','test','test.xlsx',3,4)")
        conn.execute("INSERT INTO registro VALUES (1,1,'{\"a\":1,\"b\":2}','14001','Jalisco','Guadalajara')")
        conn.execute("INSERT INTO registro VALUES (2,1,'{\"a\":3,\"b\":4}','14002','Jalisco','Zapopan')")
        conn.commit()
        conn.close()
    return path
