"""
xlsx.py — Parseador de archivos XLSX de INEGI.

Maneja dos épocas de estructura interna:
  Época 1 (2017-2019): Formato "M" — archivos CNGMD/CNGF con sub-tablas
    con encabezados multi-fila y celdas combinadas
  Época 2 (2021+): Formato temático — más consistente, aún con celdas
    combinadas en encabezados

Ambas épocas comparten:
  - Hoja "Índice" con tabla de contenido (se salta)
  - Hojas de datos numeradas (1.1, 2.3, 5.1, etc.)
  - Encabezados multi-fila (2-5 filas) con celdas combinadas
  - "Clave geográfica", "Entidad federativa", "Municipio" como primeras columnas

Estrategia de parseo:
  1. Escanear filas desde el inicio
  2. Identificar filas de título/año (se saltan)
  3. Identificar filas de encabezado (nombres de columna)
  4. Expandir forward-fill en celdas combinadas
  5. Combinar sub-encabezados con encabezados principales
  6. Leer datos desde la primera fila con códigos geográficos o valores
"""

import json
import re
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl


# ── Funciones auxiliares ──────────────────────────────────────

def _normalizar_nombre_columna(texto: str) -> Optional[str]:
    """Limpia un texto para usarlo como nombre de columna."""
    if not texto or not texto.strip():
        return None
    s = str(texto).strip()
    s = s.replace('\n', ' ').replace('\r', ' ')
    s = re.sub(r'\s+', ' ', s)
    return s.strip()


def _es_fila_vacia(row: tuple) -> bool:
    """True si todos los valores son None o string vacío."""
    return all(v is None or (isinstance(v, str) and v.strip() == '') for v in row)


def _es_header_por_contenido(texto: str, n_filled_cols: int = 0, n_unique_cols: int = 0) -> bool:
    """
    Detecta si el texto de una celda parece de encabezado de columna.
    
    Prioriza:
    1. Header keywords → header
    2. Muchas columnas ÚNICAS llenas (>= 3) → header
    3. Muchas columnas llenas pero pocas únicas → probablemente título/nota
    """
    t = texto.strip()
    if not t:
        return False
    
    # Saltar textos muy largos (títulos)
    if len(t) > 50:
        return False
    
    # Saltar títulos estándar de INEGI
    if 'INEGI' in t and 'Censo Nacional' in t:
        return False
        return False
    
    # Saltar años sueltos
    if re.match(r'^\d{4}$', t) and n_filled_cols <= 2:
        return False
    
    # Si empieza con preposición, probablemente es subtítulo no header
    if re.match(r'^(por |de |en |según |para |con |sin )', t.lower()):
        return False
    
    # Contiene saltos de línea (texto multi-línea en celda) → header
    if '\n' in t:
        return True
    
    # Palabras clave de encabezado de columna
    header_keywords = [
        'clave', 'entidad', 'municipio', 'total', 'número', 'núm',
        'año', 'sexo', 'ámbito', 'concepto', 'variable',
        'institución', 'instituci', 'sector', 'ramo',
        'función', 'función', 'prestadores', 'servicio',
        'cobertura', 'población', 'presupuesto', 'recursos',
        'personal', 'hombres', 'mujeres', 'indicador',
    ]
    t_lower = t.lower()
    if any(kw in t_lower for kw in header_keywords):
        return True
    
    # Muchas columnas ÚNICAS → casi seguro header
    # (filas de título tienen pocos valores únicos aunque ocupen muchas columnas)
    if n_unique_cols >= 3:
        return True
    
    # Muchas columnas llenas (> 5) con baja diversidad → probablemente título/nota
    if n_filled_cols >= 5 and n_unique_cols < 3:
        return False
    
    return False


# ── Detección de estructura de hoja ───────────────────────────

def _analizar_filas(ws, max_filas: int = 50) -> List[dict]:
    """Analiza las primeras filas de una hoja y retorna metadata de cada una."""
    filas = []
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i > max_filas:
            break
        
        vals = list(row)
        # Filas llenas (sin None ni vacío)
        filled = [v for v in vals if v is not None and 
                  not (isinstance(v, str) and v.strip() == '')]
        n_filled = len(filled)
        n_unique = len(set(str(v).strip() for v in filled))
        
        first = vals[0] if vals else None
        has_number = any(isinstance(v, (int, float)) for v in vals)
        
        filas.append({
            'num': i,
            'n_filled': n_filled,
            'n_unique': n_unique,
            'first': first,
            'has_number': has_number,
            'vals': vals
        })
    
    return filas


def _encontrar_estructura_hoja(ws) -> tuple:
    """
    Encuentra las filas de encabezado y la primera fila de datos.
    
    Returns:
        (fila_header_inicio, fila_header_fin, fila_datos_inicio)
        (0, 0, 0) si no se pudo detectar
    """
    filas = _analizar_filas(ws)
    
    # 1. Encontrar header_start: primera fila con contenido de encabezado
    header_start = None
    for f in filas:
        if f['first'] is None:
            continue
        first_str = str(f['first']).strip()
        # Saltar títulos muy largos
        if len(first_str) > 80:
            continue
        # Saltar años sueltos
        if re.match(r'^\d{4}$', first_str) and f['n_filled'] <= 2:
            continue
        if _es_header_por_contenido(first_str, f['n_filled'], f['n_unique']):
            header_start = f['num']
            break
    
    if header_start is None:
        return (0, 0, 0)
    
    # 2. Encontrar header_end: últimas filas consecutivas donde first=None y no hay números
    header_end = header_start
    for f in filas:
        if f['num'] <= header_start:
            continue
        # Si la fila tiene valores numéricos, es datos
        if f['has_number']:
            break
        # Si first no es None y no parece header, es datos
        if f['first'] is not None:
            first_str = str(f['first']).strip()
            # Códigos geográficos o estados
            if re.match(r'^\d{3,9}$', first_str):
                break
            if 'mexicanos' in first_str.lower():
                break
            if not _es_header_por_contenido(first_str, f['n_filled']):
                header_end = f['num']
                break
        header_end = f['num']
    
    # 3. Encontrar data_start: primera fila con números o códigos
    data_start = None
    for f in filas:
        if f['num'] <= header_end:
            continue
        # Tiene números → datos
        if f['has_number']:
            data_start = f['num']
            break
        # Primer celda con código geográfico
        if f['first'] is not None:
            first_str = str(f['first']).strip()
            if re.match(r'^\d{3,9}$', first_str):
                data_start = f['num']
                break
            if 'mexicanos' in first_str.lower():
                data_start = f['num']
                break
    
    if data_start is None or data_start <= header_end:
        return (0, 0, 0)
    
    return (header_start, header_end, data_start)


# ── Extracción de encabezados ─────────────────────────────────

def _extraer_encabezados(ws, header_start: int, header_end: int) -> Dict[int, str]:
    """
    Extrae nombres de columnas combinando múltiples filas de encabezado.
    
    Para cada columna, combina valores de todas las filas de encabezado
    en un nombre jerárquico usando ' | ' como separador.
    Aplica forward-fill para celdas combinadas.
    """
    # Leer filas de encabezado
    header_rows = []
    for i, row in enumerate(ws.iter_rows(
        min_row=header_start, max_row=header_end, values_only=True
    ), header_start):
        vals = list(row)
        header_rows.append(vals)
    
    if not header_rows:
        return {}
    
    max_col = max(len(r) for r in header_rows)
    
    # Forward-fill por fila: valores combinados se propagan a la derecha
    for row in header_rows:
        last = None
        for j in range(len(row)):
            if row[j] is not None:
                last = row[j]
            elif last is not None:
                row[j] = last  # ← forward-fill de celdas combinadas
        # Extender a max_col
        while len(row) < max_col:
            row.append(last)
    
    # Construir nombres columna por columna
    nombres = {}
    for col in range(max_col):
        partes = []
        for row in header_rows:
            val = row[col] if col < len(row) else None
            if val is not None:
                nombre = _normalizar_nombre_columna(str(val))
                if nombre:
                    # Saltar números de cuadro y notas
                    if re.match(r'^(cuadro|\d+a?\.\s?\d+|parte)$', nombre, re.IGNORECASE):
                        continue
                    # Saltar valores que son solo años
                    if re.match(r'^\d{4}$', nombre):
                        continue
                    partes.append(nombre)
        
        if partes:
            nombre_final = ' | '.join(partes) if len(partes) > 1 else partes[0]
            nombres[col] = nombre_final
    
    return nombres


# ── Parseo de hoja ────────────────────────────────────────────

def parsear_hoja(ws) -> Dict[str, Any]:
    """
    Parsea una hoja de un archivo XLSX.
    """
    nombre_hoja = ws.title.strip()
    
    header_start, header_end, data_start = _encontrar_estructura_hoja(ws)
    
    if header_start <= 0 or data_start <= 0:
        return {
            'nombre_hoja': nombre_hoja,
            'registros': [],
            'total_registros': 0,
            'errores': ['No se pudo detectar estructura']
        }
    
    nombres_col = _extraer_encabezados(ws, header_start, header_end)
    
    if not nombres_col:
        return {
            'nombre_hoja': nombre_hoja,
            'registros': [],
            'total_registros': 0,
            'errores': ['No se pudieron extraer encabezados']
        }
    
    registros = []
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if _es_fila_vacia(row):
            continue
        
        vals = list(row)
        registro = {}
        
        for col_idx, nombre in nombres_col.items():
            val = vals[col_idx] if col_idx < len(vals) else None
            if val is not None:
                val_str = str(val).strip()
                if val_str:
                    registro[nombre] = val
        
        if registro:
            registros.append(registro)
    
    return {
        'nombre_hoja': nombre_hoja,
        'registros': registros,
        'total_registros': len(registros),
        'campos': [{'nombre': n} for n in nombres_col.values()],
        'errores': []
    }


# ── Parseo de archivo XLSX completo ──────────────────────────

def parsear_xlsx(filepath: Path) -> Dict[str, Any]:
    """
    Parsea un archivo XLSX completo.
    
    1. Salta la hoja "Índice"
    2. Parsea cada hoja de datos
    3. Retorna todos los registros con metadatos
    
    Args:
        filepath: Ruta al archivo .xlsx
    
    Returns:
        Dict con nombre_archivo, censo, año, modulo, registros,
        total_registros, total_hojas, campos, errores
    """
    nombre_archivo = filepath.stem
    patron = re.match(r'(cnge|cngf|cngmd)_(\d{4})_(.+)_tabulados', nombre_archivo)
    censo = patron.group(1) if patron else 'desconocido'
    año = int(patron.group(2)) if patron else 0
    modulo = patron.group(3) if patron else 'desconocido'
    
    try:
        wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
    except Exception as e:
        return {
            'nombre_archivo': nombre_archivo,
            'censo': censo, 'año': año, 'modulo': modulo,
            'registros': [], 'total_registros': 0, 'total_hojas': 0,
            'errores': [f'No se pudo abrir XLSX: {e}']
        }
    
    todos_registros = []
    errores = []
    
    for nombre_hoja in wb.sheetnames:
        nombre_limpio = nombre_hoja.strip()
        if nombre_limpio.lower().startswith('índice') or \
           nombre_limpio.lower().startswith('indice'):
            continue
        
        ws = wb[nombre_hoja]
        try:
            res_hoja = parsear_hoja(ws)
            for reg in res_hoja['registros']:
                reg['_hoja'] = nombre_limpio
                reg['_censo'] = censo.upper()
                reg['_año'] = año
                reg['_modulo'] = modulo
            
            if res_hoja['registros']:
                todos_registros.extend(res_hoja['registros'])
            
            if res_hoja.get('errores'):
                errores.extend(f"[{nombre_limpio}] {e}" for e in res_hoja['errores'])
        except Exception as e:
            errores.append(f"[{nombre_limpio}] Error: {e}")
    
    wb.close()
    
    # Campos unificados
    campos_vistos = set()
    campos_unificados = []
    for reg in todos_registros:
        for k in reg:
            if k.startswith('_'):
                continue
            if k not in campos_vistos:
                campos_vistos.add(k)
                campos_unificados.append({'nombre': k})
    
    return {
        'nombre_archivo': nombre_archivo,
        'censo': censo, 'año': año, 'modulo': modulo,
        'registros': todos_registros,
        'total_registros': len(todos_registros),
        'total_hojas': len(set(r['_hoja'] for r in todos_registros)),
        'campos': campos_unificados,
        'errores': errores
    }
